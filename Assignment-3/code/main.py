from openpyxl import load_workbook
import pandas as pd
import numpy as np

# Load Excel file
Data = load_workbook("./../table/table.xlsx").active

#Read File
for ranofmarks in Data.iter_cols(min_col=1, max_col=1, values_only=True):
    pass
for freq in Data.iter_cols(min_col=2, max_col=2, values_only=True):
    pass
list = dict((x,y) for x,y in zip(ranofmarks,freq))

# Find probability 
def prob(key,dict):
    return "{:.3f}".format(dict[key]/dict["Total"])

def find_prob(key,dict):
    print("Probability of Marks",key,"is",prob(key,dict))

# Find required probability
find_prob('0 - 20',list)
find_prob('60 - 70',list)
find_prob('70 - 100',list)
print("Thus, Probability of Marks 60 - 100 is",float(prob('60 - 70',list)) + float(prob('70 - 100',list)))